#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
9_cleanup_duplicates.py — Mantenimiento del pipeline research_agent

Detecta y elimina artículos duplicados por DOI dentro de cada categoría
y entre categorías. Usa papers_metadata.jsonl como fuente de verdad.

Criterio de desempate para elegir cuál conservar (en orden de prioridad):
    1. Nombre limpio Crossref (todo minúsculas, sin DOI embebido en el stem)
    2. Mayor completeness (campos no vacíos en metadata)
    3. Stem alfabéticamente menor
    4. Número de línea en el JSONL

Modos de operación:
    --preview   Solo muestra duplicados detectados, sin modificar nada (defecto)
    --apply     Elimina los secundarios y reescribe papers_metadata.jsonl con backup

Ficheros leídos:
    /Volumes/research/categorias/<categoria>/metadata/papers_metadata.jsonl
    /Volumes/research/categorias/                      ← estructura completa de artefactos

Ficheros escritos (solo con --apply):
    /Volumes/research/categorias/<categoria>/metadata/
        papers_metadata.jsonl          ← reescrito sin duplicados
        papers_metadata.jsonl.bak      ← backup del original
    /Volumes/research/categorias/<categoria>/
        pdfs/<paper_id>.pdf            ← eliminado (duplicado secundario)
        tei/<paper_id>.tei.xml         ← eliminado
        md_clean/<paper_id>.clean.md   ← eliminado
        chunks/<paper_id>.jsonl        ← eliminado
        summaries/<paper_id>.summary.md← eliminado
        metadata/per_paper/<paper_id>* ← eliminado
    logs/9_cleanup_duplicates.log

Parámetros CLI:
    --preview             Solo muestra duplicados (sin modificar nada)
    --apply               Elimina duplicados y reescribe JSONL
    --category CAT        Procesa solo esta categoría (defecto: todas)
    --doi DOI             Muestra el grupo de duplicados para un DOI concreto
    --base DIR            Directorio raíz (defecto: /Volumes/research/categorias)

Dependencias:
    (solo stdlib)

Notas:
    - Ejecutar --preview periódicamente, especialmente tras ingestas masivas
      o renombrados por Crossref.
    - Tras --apply, el script indica qué categorías necesitan re-indexado FAISS:
      python3 5_build_embeddings.py --project <cat> --model bge-m3 --force
    - La detección es solo por DOI idéntico. Duplicados por título similar o
      mismo autor+año se tratan en el item 19 (detección avanzada, pendiente).
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import re
import shutil
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Optional


# ============================================================
# CONSTANTES
# ============================================================

PROJECT_ROOT = Path(__file__).parent.parent
LOG_DIR = PROJECT_ROOT / "logs"
LOG_FILE = LOG_DIR / "9_cleanup_duplicates.log"

DEFAULT_BASE = Path("/Volumes/research/categorias")

CATEGORIES = [
    "biological_gas_odor_treatment",
    "anoxic_biogas_biodesulfurization",
    "bioplastics_microplastics",
    "biogas_upgrading_biomethanation",
    "microalgae",
    "single_cell_protein",
    "advanced_oxidation_processes",
    "bioleaching_critical_materials",
]

SOURCE_FIELDS = [
    "paper_id",
    "source_file",
    "pdf_file",
    "filename",
    "file",
    "source_pdf",
    "tei_file",
]

CATEGORY_KEYWORDS = {
    "biological_gas_odor_treatment": {
        "odor", "odour", "biofilter", "biofiltration", "voc", "volatile", "gas treatment",
    },
    "anoxic_biogas_biodesulfurization": {
        "anoxic", "biodesulfurization", "biodesulphurization", "desulfurization",
        "desulphurization", "h2s", "hydrogen sulfide", "nitrate", "nitrite",
    },
    "bioplastics_microplastics": {
        "bioplastic", "bioplastics", "microplastic", "microplastics", "pla", "pbat",
        "pha", "biodegradation",
    },
    "biogas_upgrading_biomethanation": {
        "biogas upgrading", "biomethanation", "methanation", "biomethane",
        "methane", "co2", "trickle bed",
    },
    "microalgae": {
        "microalgae", "microalga", "algae", "algal", "chlorella", "spirulina",
        "scenedesmus", "photobioreactor",
    },
    "single_cell_protein": {
        "single cell protein", "scp", "microbial protein", "protein production",
        "protein-rich", "protein rich",
    },
    "advanced_oxidation_processes": {
        "advanced oxidation", "aop", "photocatalysis", "fenton", "ozonation",
        "persulfate", "hydroxyl radical",
    },
    "bioleaching_critical_materials": {
        "bioleaching", "critical materials", "critical raw materials", "rare earth",
        "lithium", "cobalt", "nickel", "e-waste", "electronic waste",
    },
}

PDF_EXTS = [".pdf", ".PDF"]

log = logging.getLogger("cleanup_duplicates")


# ============================================================
# MODELOS
# ============================================================

@dataclass(frozen=True)
class PaperRecord:
    doi: str
    category: str
    stem: str
    metadata_path: Path
    line_no: int
    raw_line: str
    data: dict[str, Any]
    source_field: str


@dataclass(frozen=True)
class DuplicateDecision:
    doi: str
    duplicate_type: str
    keep: PaperRecord
    remove: PaperRecord
    reason: str


# ============================================================
# LOGGING
# ============================================================

def setup_logging() -> logging.Logger:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("cleanup_duplicates")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    logger.addHandler(fh)

    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter("%(message)s"))
    logger.addHandler(ch)
    return logger


# ============================================================
# UTILIDADES
# ============================================================

def normalize_doi(doi: Any) -> str:
    value = str(doi or "").strip().lower()
    value = re.sub(r"^https?://(dx\.)?doi\.org/", "", value)
    value = value.rstrip("/")
    return value


def strip_known_suffixes(name: str) -> str:
    for suffix in (".tei.xml", ".clean.md", ".summary.md", ".chunks.jsonl", ".jsonl", ".xml", ".pdf"):
        if name.lower().endswith(suffix.lower()):
            return name[: -len(suffix)]
    return name


def infer_stem(rec: dict[str, Any]) -> tuple[Optional[str], str]:
    for field in SOURCE_FIELDS:
        value = rec.get(field)
        if value is None:
            continue
        if isinstance(value, str):
            value = value.strip()
        if not value:
            continue
        stem = strip_known_suffixes(Path(str(value)).name)
        if stem:
            return stem, field
    return None, ""


def nonempty_score(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, str):
        return 1 if value.strip() else 0
    if isinstance(value, (list, tuple, set, dict)):
        return 1 if value else 0
    return 1


def metadata_completeness(rec: PaperRecord) -> int:
    return sum(nonempty_score(v) for v in rec.data.values())


def category_relevance_score(rec: PaperRecord) -> int:
    text = " ".join(
        str(rec.data.get(k) or "")
        for k in ("title", "abstract", "journal", "phase", "project")
    ).lower()
    return sum(1 for kw in CATEGORY_KEYWORDS.get(rec.category, set()) if kw in text)


def file_key(stem: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", stem.lower()).strip("_")


def _is_clean_stem(stem: str) -> bool:
    """True si el stem es nombre limpio Crossref: todo minúsculas y sin DOI embebido."""
    if stem != stem.lower():
        return False
    if re.search(r"_10_\d{4,5}_", stem):
        return False
    return True


def category_dirs(base: Path, selected: Optional[list[str]]) -> list[str]:
    cats = selected or CATEGORIES
    return sorted(dict.fromkeys(cats))


def read_metadata(category: str, base: Path) -> list[PaperRecord]:
    meta_path = base / category / "metadata" / "papers_metadata.jsonl"
    if not meta_path.exists():
        log.warning("⚠ No existe metadata, salto categoría: %s", meta_path)
        return []

    records: list[PaperRecord] = []
    with meta_path.open(encoding="utf-8") as f:
        for line_no, raw in enumerate(f, start=1):
            line = raw.rstrip("\n")
            if not line.strip():
                continue
            try:
                data = json.loads(line)
            except json.JSONDecodeError as exc:
                log.warning("⚠ JSON inválido en %s:%d: %s", meta_path, line_no, exc)
                continue

            doi = normalize_doi(data.get("doi"))
            if not doi:
                continue
            stem, source_field = infer_stem(data)
            if not stem:
                log.warning("⚠ Registro con DOI sin identificador de archivo en %s:%d (%s)", meta_path, line_no, doi)
                continue
            records.append(
                PaperRecord(
                    doi=doi,
                    category=category,
                    stem=stem,
                    metadata_path=meta_path,
                    line_no=line_no,
                    raw_line=line,
                    data=data,
                    source_field=source_field,
                )
            )
    log.debug("Leídos %d registros con DOI de %s", len(records), meta_path)
    return records


def build_registry(categories: list[str], base: Path) -> dict[str, list[PaperRecord]]:
    registry: dict[str, list[PaperRecord]] = defaultdict(list)
    for category in categories:
        for rec in read_metadata(category, base):
            registry[rec.doi].append(rec)
    return dict(registry)


def choose_keep_intra(records: list[PaperRecord]) -> tuple[PaperRecord, str]:
    ordered = sorted(
        records,
        key=lambda r: (not _is_clean_stem(r.stem), -metadata_completeness(r), r.stem.lower(), r.line_no),
    )
    keep = ordered[0]
    reason = (
        f"metadata más completa ({metadata_completeness(keep)} campos no vacíos); "
        "desempate: nombre limpio Crossref > mayúsculas/DOI embebido > stem alfabético"
    )
    return keep, reason


def choose_keep_cross(records: list[PaperRecord]) -> tuple[PaperRecord, str]:
    relevance = {r: category_relevance_score(r) for r in records}
    best_score = max(relevance.values())
    best = [r for r, score in relevance.items() if score == best_score]
    if best_score > 0 and len(best) == 1:
        keep = best[0]
        return keep, f"categoría más específica por keywords ({best_score} coincidencias)"

    ordered = sorted(
        records,
        key=lambda r: (not _is_clean_stem(r.stem), r.category.lower(), -metadata_completeness(r), r.stem.lower(), r.line_no),
    )
    keep = ordered[0]
    return keep, "sin categoría específica obvia; primera por orden alfabético de categoría"


def decide_duplicates(registry: dict[str, list[PaperRecord]]) -> list[DuplicateDecision]:
    decisions: list[DuplicateDecision] = []
    for doi, records in sorted(registry.items()):
        if len(records) < 2:
            continue

        by_category: dict[str, list[PaperRecord]] = defaultdict(list)
        for rec in records:
            by_category[rec.category].append(rec)

        # Intra-categoría: decide dentro de cada categoría.
        for category, group in sorted(by_category.items()):
            if len(group) < 2:
                continue
            keep, reason = choose_keep_intra(group)
            for remove in sorted((r for r in group if r != keep), key=lambda r: (r.stem.lower(), r.line_no)):
                decisions.append(
                    DuplicateDecision(
                        doi=doi,
                        duplicate_type="intra",
                        keep=keep,
                        remove=remove,
                        reason=reason,
                    )
                )

        # Cross-categoría: usa un candidato por categoría para no mezclar
        # duplicados internos con la decisión global.
        category_representatives: list[PaperRecord] = []
        for group in by_category.values():
            if len(group) == 1:
                category_representatives.append(group[0])
            else:
                keep, _ = choose_keep_intra(group)
                category_representatives.append(keep)

        if len(category_representatives) < 2:
            continue
        keep_cross, reason = choose_keep_cross(category_representatives)
        for remove in sorted(
            (r for r in category_representatives if r != keep_cross),
            key=lambda r: (r.category.lower(), r.stem.lower(), r.line_no),
        ):
            decisions.append(
                DuplicateDecision(
                    doi=doi,
                    duplicate_type="cross",
                    keep=keep_cross,
                    remove=remove,
                    reason=reason,
                )
            )

    return dedupe_removals(decisions)


def dedupe_removals(decisions: list[DuplicateDecision]) -> list[DuplicateDecision]:
    """Evita intentar eliminar dos veces el mismo registro en casos mixtos."""
    seen: set[tuple[str, str, int, str]] = set()
    out: list[DuplicateDecision] = []
    for dec in decisions:
        key = (dec.remove.category, str(dec.remove.metadata_path), dec.remove.line_no, dec.remove.stem)
        if key in seen:
            continue
        seen.add(key)
        out.append(dec)
    return out


def pairwise_rows(decisions: list[DuplicateDecision]) -> list[tuple[str, str, str, str, str, str]]:
    rows = []
    for dec in decisions:
        rows.append((
            dec.doi,
            dec.keep.category,
            dec.keep.stem,
            dec.remove.category,
            dec.remove.stem,
            dec.duplicate_type,
        ))
    return rows


def print_table(rows: list[tuple[str, str, str, str, str, str]]) -> None:
    headers = ("DOI", "categoría_1", "archivo_1", "categoría_2", "archivo_2", "tipo")
    widths = [len(h) for h in headers]
    for row in rows:
        for i, cell in enumerate(row):
            widths[i] = min(max(widths[i], len(cell)), 52)

    def fmt_cell(cell: str, width: int) -> str:
        if len(cell) > width:
            cell = cell[: width - 1] + "…"
        return cell.ljust(width)

    print(" | ".join(fmt_cell(h, widths[i]) for i, h in enumerate(headers)))
    print("-+-".join("-" * w for w in widths))
    for row in rows:
        print(" | ".join(fmt_cell(cell, widths[i]) for i, cell in enumerate(row)))


def summarize(decisions: list[DuplicateDecision]) -> None:
    intra = [d for d in decisions if d.duplicate_type == "intra"]
    cross = [d for d in decisions if d.duplicate_type == "cross"]
    intra_categories = {d.remove.category for d in intra}
    print("")
    print(f"Duplicados intra-categoría: {len(intra)} en {len(intra_categories)} categorías")
    print(f"Duplicados cross-categoría: {len(cross)}")


def print_decision_details(decisions: list[DuplicateDecision]) -> None:
    if not decisions:
        return
    print("")
    print("Decisiones:")
    for dec in decisions:
        print(f"- DOI {dec.doi} [{dec.duplicate_type}]")
        print(f"  Conserva: {dec.keep.category}/{dec.keep.stem}")
        print(f"  Elimina : {dec.remove.category}/{dec.remove.stem}")
        print(f"  Motivo  : {dec.reason}")


def candidate_paths(project_dir: Path, stem: str) -> list[Path]:
    candidates: list[Path] = []
    candidates.extend(project_dir / "pdfs" / f"{stem}{ext}" for ext in PDF_EXTS)
    candidates.append(project_dir / "tei" / f"{stem}.tei.xml")
    candidates.append(project_dir / "tei" / f"{stem}.xml")
    candidates.append(project_dir / "md_clean" / f"{stem}.clean.md")
    candidates.append(project_dir / "summaries" / f"{stem}.summary.md")
    candidates.append(project_dir / "chunks" / f"{stem}.chunks.jsonl")
    candidates.append(project_dir / "chunks" / f"{stem}.jsonl")

    # Fallback para PDFs antiguos donde el stem del PDF difiere levemente del TEI
    # (guiones, espacios o corchetes) pero conserva la misma clave normalizada.
    pdf_dir = project_dir / "pdfs"
    if pdf_dir.exists():
        target_key = file_key(stem)
        for pdf in pdf_dir.iterdir():
            if pdf.is_file() and pdf.suffix in PDF_EXTS and file_key(pdf.stem) == target_key:
                candidates.append(pdf)

    seen: set[Path] = set()
    unique = []
    for path in candidates:
        if path not in seen:
            seen.add(path)
            unique.append(path)
    return unique


def delete_secondary_files(rec: PaperRecord, base: Path) -> list[Path]:
    project_dir = base / rec.category
    deleted: list[Path] = []
    for path in candidate_paths(project_dir, rec.stem):
        if path.exists():
            path.unlink()
            deleted.append(path)
            log.info("Eliminado: %s", path)
        else:
            log.warning("No existe, continúo: %s", path)
    return deleted


def rewrite_metadata(decisions: list[DuplicateDecision]) -> None:
    removals_by_meta: dict[Path, set[int]] = defaultdict(set)
    for dec in decisions:
        removals_by_meta[dec.remove.metadata_path].add(dec.remove.line_no)

    for meta_path, line_numbers in sorted(removals_by_meta.items(), key=lambda item: str(item[0])):
        if not meta_path.exists():
            log.warning("No existe metadata al reescribir, salto: %s", meta_path)
            continue

        backup = meta_path.with_name(meta_path.name + ".bak")
        shutil.copy2(meta_path, backup)
        log.info("Backup metadata: %s", backup)

        kept_lines: list[str] = []
        removed = 0
        with meta_path.open(encoding="utf-8") as f:
            for line_no, line in enumerate(f, start=1):
                if line_no in line_numbers:
                    removed += 1
                    continue
                kept_lines.append(line.rstrip("\n"))

        meta_path.write_text("\n".join(kept_lines) + ("\n" if kept_lines else ""), encoding="utf-8")
        log.info("Metadata reescrita: %s (%d líneas eliminadas)", meta_path, removed)


def affected_categories(decisions: Iterable[DuplicateDecision]) -> list[str]:
    return sorted({dec.remove.category for dec in decisions})


def has_bge_m3_index(category: str, base: Path) -> bool:
    embeddings_dir = base / category / "embeddings"
    if not embeddings_dir.exists():
        return False
    return any(p.is_dir() and "bge-m3" in p.name for p in embeddings_dir.iterdir())


def print_reindex_commands(categories: list[str], base: Path) -> None:
    if not categories:
        return
    print("")
    print("Categorías afectadas que necesitan re-indexar:")
    for category in categories:
        print(f"  python 5_build_embeddings.py --project {category} --model bge-m3 --force")


def apply_cleanup(decisions: list[DuplicateDecision], base: Path) -> list[Path]:
    deleted_files: list[Path] = []
    for dec in decisions:
        log.info(
            "DOI %s [%s]: conservo %s/%s; elimino %s/%s. Motivo: %s",
            dec.doi,
            dec.duplicate_type,
            dec.keep.category,
            dec.keep.stem,
            dec.remove.category,
            dec.remove.stem,
            dec.reason,
        )
        deleted_files.extend(delete_secondary_files(dec.remove, base))

    rewrite_metadata(decisions)
    return deleted_files


# ============================================================
# DETECCIÓN AVANZADA DE DUPLICADOS
# ============================================================

def normalize_title(title: str) -> str:
    if title is None:
        return ""
    text = str(title).lower()
    text = re.sub(r'[^\w\s]', '', text)
    return re.sub(r'\s+', ' ', text).strip()


def pdf_sha256(pdf_path: Path) -> Optional[str]:
    try:
        return hashlib.sha256(pdf_path.read_bytes()).hexdigest()
    except Exception:
        return None


def detect_title_duplicates(cats: list[str], base: Path) -> list[dict]:
    groups: dict[str, list[dict]] = defaultdict(list)
    for category in cats:
        meta_path = base / category / "metadata" / "papers_metadata.jsonl"
        if not meta_path.exists():
            continue
        try:
            with meta_path.open(encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        data = json.loads(line)
                    except Exception:
                        continue
                    title = data.get("title") or ""
                    norm = normalize_title(title)
                    if len(norm) < 10:
                        continue
                    stem, _ = infer_stem(data)
                    groups[norm].append({
                        "category": category,
                        "paper_id": stem or "",
                        "doi":      normalize_doi(data.get("doi") or ""),
                        "title":    title,
                        "year":     data.get("year"),
                    })
        except Exception:
            continue
    return [
        {"match_type": "title", "norm_title": norm, "papers": papers}
        for norm, papers in sorted(groups.items())
        if len(papers) >= 2
    ]


def detect_hash_duplicates(cats: list[str], base: Path) -> list[dict]:
    groups: dict[str, list[dict]] = defaultdict(list)
    for category in cats:
        pdfs_dir = base / category / "pdfs"
        if not pdfs_dir.exists():
            continue
        for pdf in sorted(pdfs_dir.iterdir()):
            if not pdf.is_file() or pdf.suffix.lower() != ".pdf":
                continue
            sha = pdf_sha256(pdf)
            if sha is None:
                continue
            groups[sha].append({
                "category": category,
                "paper_id": pdf.stem,
                "pdf_path": str(pdf),
            })
    return [
        {"match_type": "hash", "sha256": sha, "papers": papers}
        for sha, papers in sorted(groups.items())
        if len(papers) >= 2
    ]


def write_duplicate_report(
    decisions_doi: list[DuplicateDecision],
    title_dups: list[dict],
    hash_dups: list[dict],
    out_path: Path,
) -> Optional[Path]:
    try:
        import openpyxl
    except ImportError:
        print("⚠ openpyxl no instalado — no se genera el informe Excel (pip install openpyxl)")
        return None

    wb = openpyxl.Workbook()

    ws_doi = wb.active
    ws_doi.title = "DOI"
    ws_doi.append(["doi", "category_keep", "stem_keep", "category_remove", "stem_remove", "reason", "type"])
    for dec in decisions_doi:
        ws_doi.append([
            dec.doi,
            dec.keep.category,
            dec.keep.stem,
            dec.remove.category,
            dec.remove.stem,
            dec.reason,
            dec.duplicate_type,
        ])

    ws_title = wb.create_sheet("Titulo")
    ws_title.append(["norm_title", "category", "paper_id", "doi", "title", "year", "recommended_action"])
    for group in title_dups:
        for paper in group["papers"]:
            ws_title.append([
                group["norm_title"],
                paper["category"],
                paper["paper_id"],
                paper["doi"],
                paper["title"],
                paper["year"],
                "revisar_manual",
            ])

    ws_hash = wb.create_sheet("Hash")
    ws_hash.append(["sha256", "category", "paper_id", "pdf_path", "recommended_action"])
    for group in hash_dups:
        for paper in group["papers"]:
            ws_hash.append([
                group["sha256"],
                paper["category"],
                paper["paper_id"],
                paper["pdf_path"],
                "revisar_manual",
            ])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Detecta y limpia duplicados por DOI ya procesados")
    mode = ap.add_mutually_exclusive_group()
    mode.add_argument("--preview", action="store_true", help="Solo informar, no modificar nada (defecto)")
    mode.add_argument("--apply", action="store_true", help="Ejecutar la limpieza")
    ap.add_argument("--dry-run", action="store_true", help="Alias de --preview")
    ap.add_argument("--category", action="append", choices=CATEGORIES, help="Limitar a una categoría (repetible)")
    ap.add_argument("--doi", help="Resolver un DOI concreto manualmente")
    ap.add_argument("--base", default=str(DEFAULT_BASE), help=f"Directorio base (defecto: {DEFAULT_BASE})")
    return ap.parse_args()


def main() -> int:
    args = parse_args()
    setup_logging()

    base = Path(args.base)
    cats = category_dirs(base, args.category)
    doi_filter = normalize_doi(args.doi) if args.doi else ""
    apply_mode = bool(args.apply and not args.dry_run)

    log.info("Base: %s", base)
    log.info("Categorías: %s", ", ".join(cats))
    log.info("Modo: %s", "apply" if apply_mode else "preview")

    registry = build_registry(cats, base)
    if doi_filter:
        registry = {doi: records for doi, records in registry.items() if doi == doi_filter}
        if not registry:
            print(f"No se ha encontrado DOI: {doi_filter}")
            return 0

    decisions = decide_duplicates(registry)
    rows = pairwise_rows(decisions)

    if rows:
        print_table(rows)
    else:
        print("No se han encontrado duplicados por DOI.")

    print_decision_details(decisions)
    summarize(decisions)

    title_dups = detect_title_duplicates(cats, base)
    hash_dups  = detect_hash_duplicates(cats, base)
    print(f"Duplicados por título: {len(title_dups)}")
    print(f"Duplicados por hash PDF: {len(hash_dups)}")
    if title_dups or hash_dups:
        print("Ver informe Excel para revisión manual.")
    report_path = write_duplicate_report(
        decisions, title_dups, hash_dups,
        base.parent / "metadatos" / "duplicate_report.xlsx",
    )
    if report_path:
        print(report_path)

    if not apply_mode:
        print("")
        print("Usa --apply para eliminar los secundarios")
        print("No se ha modificado nada.")
        print(f"Log: {LOG_FILE}")
        return 0

    if not decisions:
        print("No hay secundarios que eliminar.")
        print(f"Log: {LOG_FILE}")
        return 0

    deleted_files = apply_cleanup(decisions, base)
    affected = affected_categories(decisions)

    print("")
    print(f"Duplicados eliminados: {len(decisions)}")
    print("Archivos eliminados:")
    if deleted_files:
        for path in deleted_files:
            print(f"  {path}")
    else:
        print("  (ninguno; los archivos secundarios ya no existían)")
    print_reindex_commands(affected, base)
    print(f"Log: {LOG_FILE}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
