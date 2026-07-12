# -*- coding: utf-8 -*-
"""
11_Articulos.py — Catálogo bibliográfico filtrable (privado y público).

Dos secciones:
  A) Tabla resumen por categoría (mismas columnas que la portada / Pendientes).
  B) Listado de artículos de una categoría (o todas) con filtros por texto
     (título/DOI/autor), año, revista y solo-con-DOI; DOI clicable, export CSV
     y métricas de la selección.

No ejecuta nada sobre el corpus: solo lee papers_metadata.jsonl por categoría.
"""

from __future__ import annotations

import importlib.util
import json
import os
import shutil
import subprocess
import sys
import time
from datetime import datetime, date
from pathlib import Path

import openpyxl
import pandas as pd
import requests
import streamlit as st

# Permitir import de app_utils.py de la carpeta padre (mismo patrón que 2_RAG.py)
STREAMLIT_APP_DIR = Path(__file__).resolve().parent.parent
SCRIPTS_DIR = STREAMLIT_APP_DIR.parent
for _d in (str(STREAMLIT_APP_DIR), str(SCRIPTS_DIR)):
    if _d not in sys.path:
        sys.path.insert(0, _d)

from app_utils import (
    CATEGORIAS_DIR, DOI_MANUAL_XLSX, METADATOS_DIR,
    check_password, is_public_app,
    list_existing_categories, get_categories_summary,
)
from utils import validation_overrides
from utils.crossref import fetch_work
from utils.metadata_validation import build_validate_cmd, issue_is_stale, year_delta_severity

# ── Helpers para DOI ──

# Email para Crossref (usa UNPAYWALL_EMAIL de config/.env como mailto educado)
_CROSSREF_MAILTO = os.getenv("UNPAYWALL_EMAIL", "")


def _fmt_authors(authors):
    """Convierte la lista de autores (dicts/strings) en texto legible."""
    if not authors:
        return ""
    if isinstance(authors, str):
        return authors
    out = []
    for a in authors:
        if isinstance(a, dict):
            fn = (a.get("forename") or "").strip()
            sn = (a.get("surname") or "").strip()
            nombre = (fn + " " + sn).strip() or (a.get("full") or "").strip()
        else:
            nombre = str(a)
        if nombre:
            out.append(nombre)
    return "; ".join(out)


def _first_author_surname(authors) -> str:
    """Apellido del primer autor (best-effort) para afinar Crossref."""
    if not authors:
        return ""
    if isinstance(authors, list):
        a = authors[0]
        if isinstance(a, dict):
            return (a.get("surname") or "").strip() or (a.get("full") or "").strip()
        return str(a).strip()
    if isinstance(authors, str):
        first = authors.split(";")[0].split(",")[0].strip()
        return first.split()[-1] if first else ""
    return ""


def crossref_suggest(title: str, mailto: str = "",
                     author: str = "", year: str = "") -> list[dict]:
    """Busca candidatos de DOI en Crossref por título bibliográfico.

    Si están disponibles, añade apellido del primer autor y año a la
    query.bibliographic para mejorar la precisión del match. El contrato
    de salida (doi/title/year/journal) no cambia.
    """
    if not title or not title.strip():
        return []
    try:
        biblio = title.strip()
        extra = " ".join(p for p in (str(author).strip(), str(year).strip()) if p)
        if extra:
            biblio = f"{biblio} {extra}"
        params = {"query.bibliographic": biblio, "rows": 3}
        if not mailto:
            mailto = _CROSSREF_MAILTO
        if mailto:
            params["mailto"] = mailto
        r = requests.get(
            "https://api.crossref.org/works",
            params=params, timeout=15,
        )
        items = r.json().get("message", {}).get("items", [])
    except Exception:
        return []
    out = []
    for it in items:
        out.append({
            "doi":     it.get("DOI", ""),
            "title":   (it.get("title") or [""])[0],
            "year":    (it.get("issued", {}).get("date-parts", [[None]])[0] or [None])[0],
            "journal": (it.get("container-title") or [""])[0],
        })
    return out


# Campos que se pueden dejar vacíos explícitamente al editar (p.ej. journal
# en un book chapter, o quitar un DOI incorrecto). title/year/authors quedan
# protegidos: un valor vacío ahí nunca se escribe (evita vaciados accidentales
# por un fallo de UI o un borrado sin querer).
CLEARABLE_FIELDS = {"journal", "doi"}


def update_metadata_fields(category: str, updates: dict) -> int:
    """Actualiza title/doi/year/authors/journal en papers_metadata.jsonl y doi_manual.xlsx.

    updates: {paper_id: {field: value, ...}} donde field ∈ {title, doi, year, authors, journal}.
    Escribe los campos presentes; journal/doi (CLEARABLE_FIELDS) se pueden
    vaciar explícitamente, el resto solo se escribe si no está vacío. Backup
    .bak previo. Devuelve el número de registros modificados en el jsonl.
    """
    if not updates:
        return 0

    meta = CATEGORIAS_DIR / category / "metadata" / "papers_metadata.jsonl"
    changed = 0
    if meta.exists():
        bak = meta.with_suffix(meta.suffix + ".bak")
        shutil.copy(meta, bak)
        lines = []
        with meta.open(encoding="utf-8") as f:
            for line in f:
                s = line.strip()
                if not s:
                    continue
                d = json.loads(s)
                pid = d.get("paper_id") or d.get("stable_id") or ""
                if pid in updates:
                    patch = updates[pid]
                    modified = False
                    for field in ("title", "doi", "year", "authors", "journal"):
                        v = patch.get(field)
                        if v is None:
                            continue
                        if v == "" or v == []:
                            if field in CLEARABLE_FIELDS and d.get(field) not in (None, "", []):
                                d[field] = "" if v == "" else []
                                modified = True
                            continue
                        d[field] = v
                        modified = True
                    if modified:
                        changed += 1
                lines.append(json.dumps(d, ensure_ascii=False))
        meta.write_text("\n".join(lines) + "\n", encoding="utf-8")

    # ── Upsert en doi_manual.xlsx (solo entradas con DOI nuevo) ──
    doi_xlsx = DOI_MANUAL_XLSX
    doi_entries = {pid: data for pid, data in updates.items() if data.get("doi")}
    if doi_entries:
        if doi_xlsx.exists():
            bak_xlsx = doi_xlsx.with_name(doi_xlsx.name + ".bak")
            shutil.copy(doi_xlsx, bak_xlsx)
        if doi_xlsx.exists():
            wb = openpyxl.load_workbook(doi_xlsx)
            ws = wb.active
            existing: dict[str, int] = {}
            if ws.max_row is not None and ws.max_row >= 2:
                for row_idx in range(2, ws.max_row + 1):
                    val = ws.cell(row=row_idx, column=1).value
                    if val is not None:
                        existing[str(val).strip()] = row_idx
            n_cols = ws.max_column or 2
            header = [ws.cell(row=1, column=c).value for c in range(1, n_cols + 1)]
            hl = [str(h).strip().lower() if h is not None else "" for h in header]
            if "fecha_inclusion" in hl:
                fecha_col = hl.index("fecha_inclusion") + 1
            else:
                fecha_col = n_cols + 1
                ws.cell(row=1, column=fecha_col).value = "fecha_inclusion"
            today = date.today().isoformat()
            for pid, udata in doi_entries.items():
                doi = udata.get("doi", "")
                if not doi:
                    continue
                nombre = pid + ".pdf"
                if nombre in existing:
                    row_idx = existing[nombre]
                    if not ws.cell(row=row_idx, column=2).value:
                        ws.cell(row=row_idx, column=2).value = doi
                else:
                    row_data = [None] * fecha_col
                    row_data[0] = nombre
                    row_data[1] = doi
                    row_data[fecha_col - 1] = today
                    ws.append(row_data)
                    existing[nombre] = ws.max_row
            wb.save(doi_xlsx)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["nombre_archivo", "doi", "fecha_inclusion"])
            today = date.today().isoformat()
            for pid, udata in doi_entries.items():
                doi = udata.get("doi", "")
                if not doi:
                    continue
                ws.append([pid + ".pdf", doi, today])
            wb.save(doi_xlsx)

    return changed


def _cell(row, field: str) -> str:
    """Lee una celda de una fila de st.data_editor como texto seguro.

    Las columnas tipadas "string" (pandas StringDtype) devuelven pd.NA —no
    None ni ""— cuando el usuario borra el contenido de la celda. str(pd.NA)
    da la cadena literal "<NA>", que un `if valor:` trata como no-vacía y
    acababa escribiéndose tal cual en papers_metadata.jsonl. Aquí se
    normaliza pd.NA/None a "" antes de comparar o guardar."""
    v = row.get(field)
    if v is None or pd.isna(v):
        return ""
    return str(v).strip()


def _parse_authors_text(text: str) -> list[dict]:
    """'Forename Surname; ...' -> [{'forename','surname'}]. Heurística: último token = apellido."""
    out = []
    for part in str(text).split(";"):
        name = part.strip()
        if not name:
            continue
        toks = name.split()
        if len(toks) >= 2:
            out.append({"forename": " ".join(toks[:-1]), "surname": toks[-1]})
        else:
            out.append({"forename": "", "surname": name})
    return out


# ── Nivel 2: validación con Crossref (sidecar + adopción) ────────────────────

def load_validation_sidecar(category: str) -> dict:
    """Lee validation_<cat>.jsonl → {paper_id: row}. {} si no existe."""
    sidecar = CATEGORIAS_DIR / category / "metadata" / f"validation_{category}.jsonl"
    out: dict = {}
    if sidecar.exists():
        with sidecar.open(encoding="utf-8") as f:
            for line in f:
                s = line.strip()
                if not s:
                    continue
                try:
                    row = json.loads(s)
                except Exception:
                    continue
                pid = row.get("paper_id") or row.get("stable_id") or ""
                if pid:
                    out[pid] = row
    return out


def _cr_authors_to_writer(cr_authors: list) -> list[dict]:
    """[{family,given}] de Crossref → [{full,forename,surname}] que espera
    update_metadata_fields (mismo formato que escribe 4_extract_metadata)."""
    out = []
    for a in cr_authors or []:
        fn = (a.get("given") or "").strip()
        sn = (a.get("family") or "").strip()
        full = (f"{fn} {sn}").strip()
        if full:
            out.append({"full": full, "forename": fn, "surname": sn})
    return out


def _load_cleanup_module():
    """Importa 9_cleanup_duplicates.py por ruta (mismo patrón que 10_Duplicados.py)."""
    key = "cleanup_dups"
    if key not in sys.modules:
        spec = importlib.util.spec_from_file_location(
            key, str(SCRIPTS_DIR / "9_cleanup_duplicates.py"))
        mod = importlib.util.module_from_spec(spec)
        sys.modules[key] = mod
        spec.loader.exec_module(mod)
    return sys.modules[key]


def _run_validator_live(category: str, crossref: bool) -> dict | None:
    """Re-ejecuta validate_metadata.py para `category` con feedback en vivo,
    mismo patrón que `execute_script_live` de 6_Mantenimiento.py (pipeline.run_step
    + status_box con salida incremental). NO reimplementa el validador."""
    import pipeline
    args = build_validate_cmd(category, crossref=crossref)
    label = f"validate_metadata.py --category {category}" + (" --crossref" if crossref else "")
    log_lines: list[str] = []
    status_box = st.status(f"Re-validando `{category}`" +
                           (" con Crossref…" if crossref else " (solo local)…"),
                           expanded=True)
    output_box = status_box.empty()
    MAX_LINES = 200

    def on_output(line: str) -> None:
        log_lines.append(line)
        output_box.code("\n".join(log_lines[-MAX_LINES:]), language="text")

    try:
        result = pipeline.run_step("validate_metadata.py", args,
                                   on_output=on_output, label=label)
    except Exception as e:
        status_box.update(label=f"✗ re-validación falló: {e}", state="error")
        st.exception(e)
        return None

    rc = result.get("returncode", -1)
    if rc == 0:
        status_box.update(label=f"✓ `{category}` re-validado", state="complete")
    else:
        status_box.update(label=f"✗ re-validación falló (rc={rc})", state="error")
    return result


def delete_papers(category: str, paper_ids: list[str]) -> dict:
    """Cuarentena reversible de cada paper + re-index FAISS de la categoría."""
    import pipeline
    from utils.constants import OLLAMA_MODEL_EMBED
    cleanup = _load_cleanup_module()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest_root = CATEGORIAS_DIR.parent / "quarantine" / "deleted" / ts
    deleted = 0
    for pid in paper_ids:
        moved, _ = cleanup.quarantine_paper(category, pid, CATEGORIAS_DIR, dest_root)
        if moved:
            deleted += 1
    if deleted:
        pipeline.run_step("5_build_embeddings.py",
                          ["--project", category, "--model", OLLAMA_MODEL_EMBED, "--force"])
    return {"deleted": deleted, "dest": str(dest_root)}


# ── Fin helpers DOI ─────────────────────────────────────────────────────────

st.set_page_config(page_title="Artículos", page_icon="📋", layout="wide")

_pwd = "PUBLIC_APP_PASSWORD" if is_public_app() else "PRIVATE_APP_PASSWORD"
if not check_password(_pwd):
    st.stop()

st.title("📋 Catálogo de artículos")

PUBLIC = is_public_app()


# ───────────────────────────────────────────────────────────────────────────
# SECCIÓN A — Tabla resumen por categoría
# ───────────────────────────────────────────────────────────────────────────

st.subheader("Resumen por categoría")


def _all_meta_mtime() -> float:
    cats = list_existing_categories()
    mtimes = []
    for cat in cats:
        p = CATEGORIAS_DIR / cat / "metadata" / "papers_metadata.jsonl"
        if p.exists():
            mtimes.append(p.stat().st_mtime)
    return max(mtimes) if mtimes else 0.0


@st.cache_data(show_spinner="Calculando resumen por categoría…")
def _summary_rows(mtime: float) -> list[dict]:
    return get_categories_summary()


summary = _summary_rows(_all_meta_mtime())
if summary:
    st.dataframe(
        pd.DataFrame(summary),
        hide_index=True,
        use_container_width=True,
        column_config={
            "PDFs": st.column_config.NumberColumn(width="small"),
            "MD limpio": st.column_config.NumberColumn(width="small"),
            "Resúmenes": st.column_config.NumberColumn(width="small"),
            "Chunks": st.column_config.NumberColumn(width="small"),
            "Metadata": st.column_config.NumberColumn(width="small"),
            "Paquetes": st.column_config.NumberColumn(width="small"),
            "Brechas": st.column_config.TextColumn(width="large"),
        },
    )
else:
    st.info("No hay categorías con PDFs en el NAS.")

st.divider()


# ───────────────────────────────────────────────────────────────────────────
# SECCIÓN B — Listado de artículos
# ───────────────────────────────────────────────────────────────────────────

st.subheader("Listado de artículos")


@st.cache_data(show_spinner="Cargando artículos…")
def load_articles(category: str, mtime: float) -> list[dict]:
    meta = CATEGORIAS_DIR / category / "metadata" / "papers_metadata.jsonl"
    rows = []
    if meta.exists():
        with meta.open(encoding="utf-8") as f:
            for line in f:
                s = line.strip()
                if not s:
                    continue
                try:
                    d = json.loads(s)
                except Exception:
                    continue
                authors_raw = d.get("authors")
                authors = _fmt_authors(authors_raw)
                first_author = _first_author_surname(authors_raw)
                refs = d.get("references")
                n_refs = len(refs) if isinstance(refs, list) else (d.get("n_references") or "")
                rows.append({
                    "category":      category,
                    "title":         d.get("title") or "",
                    "year":          d.get("year"),
                    "journal":       d.get("journal") or d.get("venue") or "",
                    "doi":           d.get("doi") or "",
                    "authors":       authors or "",
                    "first_author":  first_author,
                    "n_refs":        n_refs,
                    "quality_score": d.get("quality_score"),
                    "source_type":   d.get("source_type") or "",
                    "access_type":   d.get("access_type") or "",
                    "paper_id":      d.get("paper_id") or d.get("stable_id") or "",
                })
    return rows


def _meta_mtime(category: str) -> float:
    meta = CATEGORIAS_DIR / category / "metadata" / "papers_metadata.jsonl"
    return meta.stat().st_mtime if meta.exists() else 0.0


categories = list_existing_categories()
if not categories:
    st.info("No hay categorías en el NAS.")
    st.stop()

# ── Controles ────────────────────────────────────────────────────────────────
TODAS = "(Todas)"
sel = st.selectbox("Categoría", options=[TODAS] + categories, index=0)

if sel == TODAS:
    rows: list[dict] = []
    for cat in categories:
        rows.extend(load_articles(cat, _meta_mtime(cat)))
else:
    rows = load_articles(sel, _meta_mtime(sel))

if not rows:
    if sel == TODAS:
        st.info("No hay artículos con metadata (papers_metadata.jsonl) en ninguna categoría.")
    else:
        st.info(f"La categoría `{sel}` no tiene papers_metadata.jsonl o está vacío.")
    st.stop()

c1, c2 = st.columns([2, 1])
with c1:
    txt = st.text_input(
        "Búsqueda (título, DOI o autores)",
        value="",
        placeholder="Ej: anaerobic digestion, 10.1016/…, García",
    )
with c2:
    journals = sorted({r["journal"] for r in rows if r["journal"]})
    sel_journals = st.multiselect("Revista (opcional)", options=journals, default=[])

c3, c4 = st.columns([2, 1])
with c3:
    years = [int(r["year"]) for r in rows if isinstance(r["year"], (int, float)) or
             (isinstance(r["year"], str) and r["year"].strip().isdigit())]
    if years:
        y_min, y_max = min(years), max(years)
        usar_year = st.checkbox("Filtrar por año", value=False)
        if usar_year and y_min < y_max:
            ys, ye = st.slider("Rango de años", y_min, y_max, (y_min, y_max))
        elif usar_year:
            ys = ye = y_min
        else:
            ys = ye = None
    else:
        ys = ye = None
with c4:
    doi_filter = st.radio(
        "DOI", options=["Todos", "Con DOI", "Sin DOI"],
        index=0, horizontal=True, key="doi_filter",
    )
    solo_doi = (doi_filter == "Con DOI")
    solo_sin_doi = (doi_filter == "Sin DOI")


# ── Aplicar filtros en cascada ───────────────────────────────────────────────
def _year_int(v):
    if isinstance(v, (int, float)):
        return int(v)
    if isinstance(v, str) and v.strip().isdigit():
        return int(v.strip())
    return None


filtered = rows
if txt.strip():
    q = txt.strip().lower()
    filtered = [
        r for r in filtered
        if q in str(r["title"]).lower()
        or q in str(r["doi"]).lower()
        or q in str(r["authors"]).lower()
    ]
if sel_journals:
    jset = set(sel_journals)
    filtered = [r for r in filtered if r["journal"] in jset]
if ys is not None and ye is not None:
    filtered = [
        r for r in filtered
        if (_y := _year_int(r["year"])) is not None and ys <= _y <= ye
    ]
if solo_sin_doi:
    filtered = [r for r in filtered if not str(r["doi"]).strip()]
elif solo_doi:
    filtered = [r for r in filtered if str(r["doi"]).strip()]

st.caption(f"**{len(filtered)}** artículo(s) tras filtrar (de {len(rows)} cargados).")

if not filtered:
    st.warning("Sin resultados con esos filtros. Prueba aflojando los filtros.")
    st.stop()

# ── Tira de métricas ─────────────────────────────────────────────────────────
n_art = len(filtered)
n_doi = sum(1 for r in filtered if str(r["doi"]).strip())
pct_doi = round(100 * n_doi / n_art) if n_art else 0
f_years = [y for r in filtered if (y := _year_int(r["year"])) is not None]
qscores = [float(r["quality_score"]) for r in filtered
           if isinstance(r["quality_score"], (int, float))]

m1, m2, m3, m4 = st.columns(4)
m1.metric("Artículos", n_art)
m2.metric("% con DOI", f"{pct_doi}%")
m3.metric("Años", f"{min(f_years)}–{max(f_years)}" if f_years else "—")
m4.metric("Calidad media", f"{sum(qscores) / len(qscores):.2f}" if qscores else "—")

# ── Tabla ────────────────────────────────────────────────────────────────────
df = pd.DataFrame(filtered)
df["doi_url"] = df["doi"].apply(
    lambda d: f"https://doi.org/{d}" if str(d).strip() else "")

if PUBLIC:
    cols = ["title", "year", "journal", "doi_url"]
else:
    cols = ["title", "year", "journal", "doi_url", "authors", "n_refs",
            "quality_score", "source_type", "access_type", "paper_id"]

col_config = {
    "title":         st.column_config.TextColumn("Título", width="large"),
    "year":          st.column_config.NumberColumn("Año", width="small"),
    "journal":       st.column_config.TextColumn("Revista"),
    "doi_url":       st.column_config.LinkColumn(
        "DOI", display_text=r"https://doi\.org/(.+)"),
    "authors":       st.column_config.TextColumn("Autores"),
    "n_refs":        st.column_config.NumberColumn("Refs", width="small"),
    "quality_score": st.column_config.NumberColumn("Calidad", width="small"),
    "source_type":   st.column_config.TextColumn("Origen"),
    "access_type":   st.column_config.TextColumn("Acceso"),
    "paper_id":      st.column_config.TextColumn("paper_id"),
}

st.dataframe(
    df[cols],
    hide_index=True,
    use_container_width=True,
    column_config=col_config,
)

# ── Export CSV de la vista filtrada ──────────────────────────────────────────
_today = datetime.now().strftime("%Y-%m-%d")
_slug = "todas" if sel == TODAS else sel
csv_bytes = df[cols].to_csv(index=False).encode("utf-8-sig")
st.download_button(
    "⬇️ Exportar CSV (vista filtrada)",
    data=csv_bytes,
    file_name=f"articulos_{_slug}_{_today}.csv",
    mime="text/csv",
)


# ───────────────────────────────────────────────────────────────────────────
# ✏️ Editar / 🗑 eliminar artículos (SOLO INSTANCIA PRIVADA)
# ───────────────────────────────────────────────────────────────────────────
if not PUBLIC:
    st.divider()
    st.subheader("✏️ Editar / 🗑 eliminar artículos")
    if sel == TODAS:
        st.info("Selecciona una categoría concreta arriba para editar o eliminar.")
    else:
        edit_rows = load_articles(sel, _meta_mtime(sel))
        if not edit_rows:
            st.info("Sin artículos en esta categoría.")
        else:
            nonce = st.session_state.get("art_editor_nonce", 0)

            # ── Filtros del editor ──
            fcol1, fcol2 = st.columns([2, 1])
            with fcol1:
                q_ed = st.text_input(
                    "Buscar (título, DOI o autor)", value="",
                    key=f"art_edit_search_{sel}",
                    placeholder="Ej: siloxane, 10.1016/…, López",
                )
            with fcol2:
                mostrar = st.radio(
                    "Mostrar", ["Todos", "Sin DOI", "Sin año", "Sin autores", "Incompletos"],
                    index=0, key=f"art_edit_show_{sel}",
                )

            def _incompleto(r):
                return (not str(r.get("doi") or "").strip()
                        or r.get("year") in (None, "")
                        or not str(r.get("authors") or "").strip())

            if q_ed.strip():
                qq = q_ed.strip().lower()
                edit_rows = [r for r in edit_rows
                             if qq in str(r.get("title", "")).lower()
                             or qq in str(r.get("doi", "")).lower()
                             or qq in str(r.get("authors", "")).lower()]
            if mostrar == "Sin DOI":
                edit_rows = [r for r in edit_rows if not str(r.get("doi") or "").strip()]
            elif mostrar == "Sin año":
                edit_rows = [r for r in edit_rows if r.get("year") in (None, "")]
            elif mostrar == "Sin autores":
                edit_rows = [r for r in edit_rows if not str(r.get("authors") or "").strip()]
            elif mostrar == "Incompletos":
                edit_rows = [r for r in edit_rows if _incompleto(r)]

            st.caption(f"{len(edit_rows)} artículo(s) en el editor.")

            sug_key = f"doi_suggestions_{sel}"
            if not edit_rows:
                st.info("Sin artículos con esos filtros.")
            else:
                if st.button("🔎 Sugerir DOIs faltantes (Crossref)", use_container_width=True):
                    sin = [r for r in edit_rows if not str(r["doi"]).strip()]
                    with st.spinner(f"Consultando Crossref para {len(sin)} artículo(s)…"):
                        suggestions = {}
                        for i, r in enumerate(sin):
                            if r.get("title"):
                                cands = crossref_suggest(r["title"], author=r.get("first_author", ""),
                                                         year=str(r.get("year") or ""))
                                if cands:
                                    suggestions[r["paper_id"]] = cands[0].get("doi", "")
                            if i < len(sin) - 1:
                                time.sleep(0.3)
                    st.session_state[sug_key] = suggestions
                    st.session_state["art_editor_nonce"] = nonce + 1
                    st.success(f"Crossref sugirió {len(suggestions)} DOI(s).")
                    st.rerun()
                suggestions = st.session_state.get(sug_key, {})

                editor_rows = []
                for r in edit_rows:
                    pid = r["paper_id"]
                    editor_rows.append({
                        "_sel": False,
                        "paper_id": pid,
                        "title": r.get("title", ""),
                        "year": "" if r.get("year") in (None, "") else str(r.get("year")),
                        "authors": r.get("authors", ""),
                        "journal": r.get("journal", ""),
                        "doi": str(r.get("doi") or "") or suggestions.get(pid, ""),
                    })
                df_ed = pd.DataFrame(editor_rows).astype(
                    {"title": "string", "year": "string", "authors": "string",
                     "journal": "string", "doi": "string"})
                col_cfg = {
                    "_sel": st.column_config.CheckboxColumn("🗑", width="small", default=False),
                    "paper_id": st.column_config.TextColumn("paper_id", disabled=True),
                    "title": st.column_config.TextColumn("Título", width="large"),
                    "year": st.column_config.TextColumn("Año", width="small"),
                    "authors": st.column_config.TextColumn("Autores (sep. ';')", width="large"),
                    "journal": st.column_config.TextColumn("Revista", width="medium"),
                    "doi": st.column_config.TextColumn("DOI", width="medium"),
                }
                edited = st.data_editor(
                    df_ed, column_config=col_cfg,
                    column_order=["_sel", "paper_id", "title", "year", "authors", "journal", "doi"],
                    hide_index=True, use_container_width=True, num_rows="fixed",
                    key=f"art_editor_{sel}_{nonce}_{abs(hash((q_ed, mostrar)))}",
                )
                orig = {r["paper_id"]: r for r in edit_rows}

                cga, cgb = st.columns(2)
                with cga:
                    if st.button("💾 Guardar cambios", type="primary", use_container_width=True):
                        updates = {}
                        for _, row in edited.iterrows():
                            pid = row["paper_id"]; o = orig.get(pid, {})
                            u = {}
                            ti_v = _cell(row, "title")
                            if ti_v != str(o.get("title") or "").strip():
                                u["title"] = ti_v
                            doi_v = _cell(row, "doi")
                            if doi_v != str(o.get("doi") or "").strip():
                                u["doi"] = doi_v
                            yr_v = _cell(row, "year")
                            if yr_v != ("" if o.get("year") in (None, "") else str(o.get("year"))):
                                u["year"] = int(yr_v) if yr_v.isdigit() else (yr_v or None)
                            au_v = _cell(row, "authors")
                            if au_v != str(o.get("authors") or "").strip():
                                u["authors"] = _parse_authors_text(au_v)
                            jr_v = _cell(row, "journal")
                            if jr_v != str(o.get("journal") or "").strip():
                                u["journal"] = jr_v
                            if u:
                                updates[pid] = u
                        if not updates:
                            st.info("No hay cambios que guardar.")
                        else:
                            n = update_metadata_fields(sel, updates)
                            st.success(f"✓ {n} registro(s) actualizado(s) (backup .bak).")
                            st.session_state.pop(sug_key, None)
                            st.session_state["art_editor_nonce"] = nonce + 1
                            load_articles.clear(); _summary_rows.clear()
                            st.rerun()
                with cgb:
                    to_delete = [row["paper_id"] for _, row in edited.iterrows() if row.get("_sel")]
                    st.caption(f"{len(to_delete)} marcado(s) para eliminar.")
                    confirm = st.checkbox("Confirmo eliminar (reversible → quarantine/)",
                                          key=f"confirm_del_{sel}")
                    if st.button("🗑 Eliminar seleccionados", use_container_width=True,
                                 disabled=not (to_delete and confirm)):
                        with st.spinner(f"Eliminando {len(to_delete)} y re-indexando…"):
                            res = delete_papers(sel, to_delete)
                        st.success(f"✓ {res['deleted']} eliminado(s) → {res['dest']}. FAISS re-indexado.")
                        st.session_state["art_editor_nonce"] = nonce + 1
                        load_articles.clear(); _summary_rows.clear()
                        st.rerun()


# ───────────────────────────────────────────────────────────────────────────
# 🔬 Validación de metadata (Crossref, Nivel 2) — SOLO INSTANCIA PRIVADA
# ───────────────────────────────────────────────────────────────────────────
if not PUBLIC:
    st.divider()
    st.subheader("🔬 Validación de metadata (Crossref)")
    st.caption("Crossref es la mejor fuente, no infalible; revisa antes de adoptar. "
               "Toda escritura pasa por el editor (backup `.bak`).")

    if sel == TODAS:
        st.info("Selecciona una categoría concreta arriba para revisar discrepancias.")
    else:
        sidecar = load_validation_sidecar(sel)
        if not sidecar:
            st.info(
                f"No hay `validation_{sel}.jsonl`. Genéralo en pciq22 con "
                f"`python3 scripts/validate_metadata.py --category {sel} --crossref`.")
        else:
            # Campos marcados "verificado, no volver a sugerir" (persiste en
            # metadatos/validation_overrides.csv; filtra listado por-campo Y
            # las adopciones masivas de abajo).
            dismissed = validation_overrides.dismissed_set(sel)

            solo_disc = st.checkbox(
                "⚠ Solo con discrepancias (del sidecar)", value=True,
                key=f"solo_disc_{sel}")

            # ── (i) Listado por-campo del sidecar (adopción individual) ──
            all_rows = load_articles(sel, _meta_mtime(sel))
            by_pid = {r["paper_id"]: r for r in all_rows}
            pids = list(sidecar.keys()) if solo_disc else [r["paper_id"] for r in all_rows]

            col_cnt, col_cr, col_rv = st.columns([4, 3, 2])
            col_cnt.caption(f"{len(sidecar)} paper(s) marcados para revisar en esta categoría.")
            incluir_crossref = col_cr.checkbox(
                "Incluir Crossref (más lento)", value=True, key=f"cr_toggle_{sel}")
            revalidar = col_rv.button("🔄 Re-validar esta categoría", key=f"revalidate_{sel}")
            st.caption("Re-ejecuta el validador y refresca el sidecar de esta categoría. "
                       "Con Crossref tarda ~1 min.")
            if revalidar:
                result = _run_validator_live(sel, incluir_crossref)
                if result and result.get("returncode") == 0:
                    st.rerun()
                elif result:
                    tail = "\n".join(result.get("output", [])[-15:])
                    st.error(f"La re-validación falló (rc={result.get('returncode')}):\n\n{tail}")

            n_resueltos = 0  # issues de la foto ya adoptados (vigente == sugerido)

            for pid in pids:
                row = sidecar.get(pid)
                title_txt = (by_pid.get(pid, {}).get("title")
                             or (row or {}).get("title") or pid)[:80]
                if not row:
                    # Checkbox OFF: paper de la categoría sin issues en el sidecar.
                    with st.expander(f"`{pid}` — {title_txt}"):
                        st.caption("Sin issues en el último pase del validador.")
                    continue

                local_issues = row.get("issues") or []
                cr_block = row.get("crossref")  # None → pase sin --crossref
                cr_issues = (cr_block or {}).get("issues") or []
                # Adoptables campo a campo: title/journal/year con sugerencia.
                # year puede venir por duplicado (fuente paper_id y crossref):
                # cada issue lleva su propio botón etiquetado con la fuente.
                suggests = [i for i in cr_issues
                            if i.get("suggested")
                            and i.get("kind") in ("mismatch", "recover", "fill")
                            and i.get("field") in ("title", "journal", "year")]
                authors_iss = [i for i in cr_issues
                               if i.get("code") == "authors_mismatch"]
                # DOI correcto pero no resuelto en Crossref (404/miss) — no es
                # una sugerencia de valor, solo un aviso informativo que
                # también se puede "mantener" (dejar de mostrar).
                doi_miss = bool(cr_block) and not cr_block.get("fetched", True)

                # Verificados: el usuario ya revisó y descartó estas
                # sugerencias — no volver a mostrarlas (utils.validation_overrides).
                suggests = [i for i in suggests
                            if (pid, i.get("field")) not in dismissed]
                local_issues = [i for i in local_issues
                                if (pid, i.get("field")) not in dismissed]
                if (pid, "authors") in dismissed:
                    authors_iss = []
                if (pid, "doi") in dismissed:
                    doi_miss = False

                # Desaparición en vivo: omite issues que ya no aplican al
                # registro vigente (sugerencia adoptada / campo cambiado). El
                # sidecar en disco no se toca — es la foto del último pase.
                cur_rec = by_pid.get(pid, {})
                n_antes = len(suggests) + len(local_issues)
                suggests = [i for i in suggests
                            if not issue_is_stale(i, cur_rec)]
                local_issues = [i for i in local_issues
                                if not issue_is_stale(i, cur_rec, row)]
                stale_aqui = n_antes - len(suggests) - len(local_issues)
                n_resueltos += stale_aqui
                if not suggests and not local_issues and not authors_iss and not doi_miss:
                    if stale_aqui:
                        with st.expander(f"✓ `{pid}` — {title_txt} "
                                         f"(resuelto en esta sesión)"):
                            st.caption("Los campos adoptados ya coinciden con la "
                                       "sugerencia; re-ejecuta el validador para "
                                       "refrescar el sidecar.")
                    continue

                label = (f"`{pid}` — {title_txt}  "
                         f"({len(local_issues)} local(es) · "
                         f"{len(suggests)} sugerencia(s) Crossref)")
                with st.expander(label):
                    for k, iss in enumerate(suggests):
                        field = iss["field"]
                        src = iss.get("suggested_source", "crossref")
                        st.markdown(
                            f"**{field}** ({iss['code']}, fuente {src})  \n"
                            f"- actual: `{iss.get('stored')}`  \n"
                            f"- sugerido: `{iss.get('suggested')}`")
                        cA, cB = st.columns(2)
                        if cA.button(f"Adoptar {field} ({src})",
                                     key=f"adopt_{sel}_{pid}_{field}_{src}_{k}"):
                            val = iss["suggested"]
                            if field == "year":
                                val = int(val) if str(val).isdigit() else val
                            update_metadata_fields(sel, {pid: {field: val}})
                            st.success(f"✓ {field} adoptado (backup .bak).")
                            load_articles.clear(); _summary_rows.clear()
                            st.rerun()
                        if cB.button(f"✋ Mantener {field} (no volver a sugerir)",
                                     key=f"keep_{sel}_{pid}_{field}_{src}_{k}"):
                            validation_overrides.dismiss(
                                sel, pid, field,
                                note=f"rechazada sugerencia {src}: {iss.get('suggested')}")
                            st.success(f"✓ {field} marcado como verificado — "
                                       "no se volverá a sugerir.")
                            st.rerun()

                    if authors_iss:
                        st.caption("↳ Autores difieren de Crossref — se reparan con "
                                   "la adopción masiva de abajo, no campo a campo.")
                        if st.button("✋ Mantener autores (no volver a sugerir)",
                                     key=f"keep_authors_field_{sel}_{pid}"):
                            validation_overrides.dismiss(
                                sel, pid, "authors",
                                note="rechazada propuesta de adopción masiva de autores")
                            st.success("✓ autores marcados como verificados.")
                            st.rerun()
                    if doi_miss:
                        st.caption("↳ DOI no resuelto en Crossref (miss/404) — "
                                   "puede ser un DOI correcto que Crossref no "
                                   "indexa (p.ej. capítulo de libro).")
                        if st.button("✋ Mantener DOI (no volver a marcar como no resuelto)",
                                     key=f"keep_doi_miss_{sel}_{pid}"):
                            validation_overrides.dismiss(
                                sel, pid, "doi",
                                note="DOI confirmado correcto pese a miss/404 en Crossref")
                            st.success("✓ DOI marcado como verificado — no se "
                                       "volverá a avisar de este miss.")
                            st.rerun()

                    if local_issues:
                        st.markdown("**Diagnóstico local (Nivel 1)** — solo lectura:")
                        for iss in local_issues:
                            st.markdown(f"- `{iss.get('code')}` "
                                        f"({iss.get('severity')}): {iss.get('msg')}")

                    if cr_block is None:
                        st.caption("ℹ Re-ejecuta el validador con `--crossref` "
                                   "para ver sugerencias de Crossref.")
                    elif not suggests and not authors_iss and not local_issues and not doi_miss:
                        st.caption("Sin discrepancias accionables en este pase.")

            if n_resueltos:
                st.caption(f"{n_resueltos} issue(s) resueltos en esta sesión; "
                           "re-ejecuta el validador para refrescar el sidecar.")

            # ── (ii) Adopción MASIVA de autores por categoría (preview→confirmar) ──
            st.markdown("---")
            st.markdown("**Adopción masiva de autores de Crossref** "
                        "(todos los papers con DOI de esta categoría)")
            prev_key = f"authors_preview_{sel}"

            if st.button("👁 Previsualizar adopción de autores", key=f"prev_btn_{sel}"):
                with st.spinner("Consultando Crossref por DOI…"):
                    preview = []
                    for r in all_rows:
                        pid = r["paper_id"]
                        if (pid, "authors") in dismissed:
                            continue
                        doi = str(r.get("doi") or "").strip()
                        if not doi:
                            preview.append({"paper_id": pid, "estado": "sin DOI",
                                            "actual": r.get("authors", ""),
                                            "crossref": "", "cambia": False})
                            continue
                        work = fetch_work(doi)
                        cr_auth = _cr_authors_to_writer((work or {}).get("authors") or [])
                        if not cr_auth:
                            preview.append({"paper_id": pid, "estado": "sin fuente",
                                            "actual": r.get("authors", ""),
                                            "crossref": "", "cambia": False})
                            continue
                        cr_txt = "; ".join(a["full"] for a in cr_auth)
                        preview.append({
                            "paper_id": pid, "estado": "ok",
                            "actual": r.get("authors", ""), "crossref": cr_txt,
                            "cambia": cr_txt.strip() != str(r.get("authors") or "").strip(),
                            "_authors": cr_auth})
                st.session_state[prev_key] = preview
                st.rerun()

            preview = st.session_state.get(prev_key)
            if preview:
                n_cambia = sum(1 for p in preview if p["cambia"])
                n_sin = sum(1 for p in preview if p["estado"] != "ok")
                st.caption(f"**{n_cambia}** paper(s) se actualizarán · "
                           f"{n_sin} sin fuente/DOI (se saltan).")
                st.dataframe(
                    pd.DataFrame([{k: v for k, v in p.items() if not k.startswith("_")}
                                  for p in preview]),
                    hide_index=True, use_container_width=True)

                cambia_rows = [p for p in preview if p["cambia"]]
                if cambia_rows:
                    st.caption("Descarta aquí los que NO quieras aplicar "
                               "(no se volverán a sugerir):")
                    for p in cambia_rows:
                        pid_a = p["paper_id"]
                        cc1, cc2 = st.columns([6, 2])
                        cc1.markdown(f"`{pid_a}` — actual: `{p['actual']}` "
                                     f"→ Crossref: `{p['crossref']}`")
                        if cc2.button("✋ Mantener autores",
                                      key=f"keep_authors_mass_{sel}_{pid_a}"):
                            validation_overrides.dismiss(
                                sel, pid_a, "authors",
                                note=f"rechazada propuesta masiva: {p['crossref']}")
                            st.session_state[prev_key] = [
                                x for x in preview if x["paper_id"] != pid_a]
                            st.rerun()

                conf = st.checkbox("Confirmar adopción de autores",
                                   key=f"conf_authors_{sel}")
                if st.button("✅ Confirmar adopción", key=f"do_authors_{sel}",
                             disabled=not conf):
                    updates = {p["paper_id"]: {"authors": p["_authors"]}
                               for p in preview if p["cambia"] and p.get("_authors")}
                    n = update_metadata_fields(sel, updates)
                    st.success(f"✓ {n} paper(s) con autores actualizados (backup .bak).")
                    st.session_state.pop(prev_key, None)
                    load_articles.clear(); _summary_rows.clear()
                    st.rerun()

            # ── (iii) Adopción MASIVA de año por categoría (preview→confirmar).
            # SOLO entra el lote severity "low" (|delta|==1, online-temprano vs
            # print): los |delta|>=2 se muestran aparte y se adoptan uno a uno
            # en el listado por-campo de arriba (caso passalacqua: print alto
            # defendible pero no automático).
            st.markdown("---")
            st.markdown("**Adopción masiva de año de Crossref** "
                        "(papers con DOI de esta categoría; año canónico = "
                        "published-print; solo desfases ±1 — los saltos "
                        "grandes se revisan uno a uno)")
            year_prev_key = f"year_preview_{sel}"

            if st.button("👁 Previsualizar adopción de año",
                         key=f"year_prev_btn_{sel}"):
                with st.spinner("Consultando Crossref por DOI…"):
                    ymass, yreview, yskip = [], [], []
                    for r in all_rows:
                        pid = r["paper_id"]
                        if (pid, "year") in dismissed:
                            continue
                        doi = str(r.get("doi") or "").strip()
                        if not doi:
                            yskip.append({"paper_id": pid, "motivo": "sin DOI"})
                            continue
                        work = fetch_work(doi)
                        if not work:
                            yskip.append({"paper_id": pid,
                                          "motivo": "miss Crossref"})
                            continue
                        cr_year = work.get("year")
                        if not cr_year:
                            yskip.append({"paper_id": pid,
                                          "motivo": "sin año Crossref"})
                            continue
                        try:
                            local = int(r.get("year"))
                        except (TypeError, ValueError):
                            local = None
                        if local == cr_year:
                            continue  # ya coincide: nada que adoptar
                        delta = (cr_year - local) if local is not None else None
                        entry = {"paper_id": pid, "año actual": local,
                                 "año Crossref": int(cr_year), "delta": delta}
                        if year_delta_severity(delta) == "low":
                            ymass.append(entry)
                        else:
                            yreview.append(entry)
                    ymass.sort(key=lambda p: p["paper_id"])
                    yreview.sort(key=lambda p: p["paper_id"])
                    st.session_state[year_prev_key] = {
                        "mass": ymass, "review": yreview, "skip": yskip}
                st.rerun()

            ydata = st.session_state.get(year_prev_key)
            if ydata:
                ymass = ydata.get("mass", [])
                yreview = ydata.get("review", [])
                yskip = ydata.get("skip", [])
                st.caption(f"**{len(ymass)}** paper(s) +1 se actualizarán "
                           f"(adopción masiva) · **{len(yreview)}** con salto "
                           f"grande a revisar aparte · {len(yskip)} se saltan "
                           f"(sin DOI / miss / sin año).")
                if ymass:
                    st.caption("Descarta aquí los que NO quieras aplicar "
                               "(no se volverán a sugerir):")
                    for p in ymass:
                        pid_m = p["paper_id"]
                        mc1, mc2, mc3, mc4 = st.columns([4, 2, 3, 2])
                        mc1.markdown(f"`{pid_m}`")
                        mc2.markdown(f"actual: `{p['año actual']}`")
                        mc3.markdown(f"Crossref: `{p['año Crossref']}` "
                                     f"(Δ {p['delta']})")
                        if mc4.button("✋ Mantener año",
                                      key=f"keep_year_mass_{sel}_{pid_m}"):
                            validation_overrides.dismiss(
                                sel, pid_m, "year",
                                note=f"rechazada propuesta masiva ±1: {p['año Crossref']}")
                            ydata["mass"] = [x for x in ymass if x["paper_id"] != pid_m]
                            st.session_state[year_prev_key] = ydata
                            st.rerun()
                    conf_y = st.checkbox("Confirmar adopción de año (solo ±1)",
                                         key=f"conf_year_{sel}")
                    if st.button("✅ Confirmar adopción de año",
                                 key=f"do_year_{sel}", disabled=not conf_y):
                        updates = {p["paper_id"]: {"year": p["año Crossref"]}
                                   for p in ymass}
                        n = update_metadata_fields(sel, updates)
                        st.success(f"✓ {n} paper(s) con año actualizado "
                                   f"(backup .bak).")
                        st.session_state.pop(year_prev_key, None)
                        load_articles.clear(); _summary_rows.clear()
                        st.rerun()
                elif not yreview:
                    st.caption("Todos los años con fuente Crossref ya coinciden.")
                if yreview:
                    st.markdown("**Revisar individualmente** (no incluidos en "
                                "la adopción masiva)")
                    st.caption("Saltos |Δ|≥2: corrupción probable o caso "
                               "editorial dudoso — compara actual vs sugerido "
                               "antes de adoptar (el passalacqua es criterio "
                               "del usuario). Adopción individual in situ, "
                               "misma escritura que el listado por-campo.")
                    # Desaparición en vivo: fuera los que ya coinciden con
                    # Crossref (adoptados aquí o en el listado por-campo).
                    vivos = [p for p in yreview if not issue_is_stale(
                        {"field": "year", "suggested": p["año Crossref"]},
                        by_pid.get(p["paper_id"], {}))]
                    if not vivos:
                        st.caption("✓ Todos los saltos grandes están "
                                   "resueltos en esta sesión.")
                    for p in vivos:
                        pid_y = p["paper_id"]
                        c1, c2, c3, c4, c5 = st.columns([4, 2, 3, 2, 2])
                        c1.markdown(f"`{pid_y}`")
                        c2.markdown(f"actual: `{p['año actual']}`")
                        c3.markdown(f"Crossref: `{p['año Crossref']}` "
                                    f"(Δ {p['delta']})")
                        if c4.button("Adoptar año",
                                     key=f"adopt_big_{sel}_{pid_y}"):
                            update_metadata_fields(
                                sel, {pid_y: {"year": p["año Crossref"]}})
                            st.success("✓ año adoptado (backup .bak).")
                            load_articles.clear(); _summary_rows.clear()
                            st.rerun()
                        if c5.button("✋ Mantener año",
                                     key=f"keep_year_review_{sel}_{pid_y}"):
                            validation_overrides.dismiss(
                                sel, pid_y, "year",
                                note=f"rechazada propuesta Δ{p['delta']}: {p['año Crossref']}")
                            ydata["review"] = [x for x in yreview if x["paper_id"] != pid_y]
                            st.session_state[year_prev_key] = ydata
                            st.rerun()
                if yskip:
                    with st.expander(f"Saltados ({len(yskip)})"):
                        st.caption("Sin DOI / miss Crossref / sin año — marca "
                                   "'no reintentar' si ya sabes que no hay nada "
                                   "que hacer con este paper.")
                        for p in yskip:
                            pid_s = p["paper_id"]
                            sc1, sc2, sc3 = st.columns([5, 3, 2])
                            sc1.markdown(f"`{pid_s}`")
                            sc2.markdown(p["motivo"])
                            if sc3.button("🚫 No reintentar año",
                                          key=f"keep_year_skip_{sel}_{pid_s}"):
                                validation_overrides.dismiss(
                                    sel, pid_s, "year", note=f"saltado: {p['motivo']}")
                                ydata["skip"] = [x for x in yskip if x["paper_id"] != pid_s]
                                st.session_state[year_prev_key] = ydata
                                st.rerun()

            # ── (iv) Campos marcados como verificados — revisar/revertir ──
            st.markdown("---")
            overrides_df = validation_overrides.list_dismissed(sel)
            with st.expander(f"🔓 Campos marcados como verificados "
                             f"({len(overrides_df)})"):
                if overrides_df.empty:
                    st.caption("Ninguno en esta categoría.")
                else:
                    for _, r in overrides_df.iterrows():
                        oc1, oc2, oc3, oc4 = st.columns([3, 2, 4, 2])
                        oc1.markdown(f"`{r['paper_id']}`")
                        oc2.markdown(f"**{r['field']}**")
                        oc3.caption(r.get("note", "") or "—")
                        if oc4.button("↩ Revertir",
                                      key=f"undismiss_{sel}_{r['paper_id']}_{r['field']}"):
                            validation_overrides.undismiss(sel, r["paper_id"], r["field"])
                            st.rerun()
