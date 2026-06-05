#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
1_rename_papers_by_doi.py — Paso 1 del pipeline research_agent

Renombra artículos PDF científicos usando el DOI y los metadatos de Crossref.

Formato de nombre resultante:
    año_apellidoprimerautor_resumentitulo.pdf
    Ejemplo: 2024_garcia_anaerobic_biodegradation_pla_pbat_sludge.pdf

Posición en el pipeline:
    inbox/ → 1_rename_papers_by_doi.py → inbox/ (renombrado) → 2_screen_pdfs.py

Estrategia de extracción de DOI (en orden):
    1. Metadatos PDF (campo /doi, /Subject, /Keywords)
    2. Regex sobre el texto de las primeras 3 páginas
    3. Lookup en doi_manual.xlsx si el PDF está registrado manualmente

Ficheros leídos:
    <folder>/*.pdf                              ← PDFs a renombrar
    /Volumes/research/metadatos/doi_manual.xlsx ← DOIs manuales (opcional)
    config/.env                                 ← CROSSREF_EMAIL (opcional)

Ficheros escritos:
    <folder>/*.pdf                              ← PDFs renombrados (con --apply)
    /Volumes/research/fallidos/                 ← PDFs sin DOI (con --move-failed)
    <csv>                                       ← Informe CSV del proceso
    /Volumes/research/metadatos/doi_manual.xlsx ← Actualizado con DOIs no encontrados

Parámetros CLI:
    --folder DIR          Carpeta con los PDFs (defecto: /Volumes/research/inbox)
    --csv PATH            Ruta del CSV de salida
    --apply               Ejecuta el renombrado (sin esta flag: solo preview)
    --move-failed         Mueve los PDFs sin DOI a /Volumes/research/fallidos/
    --max-words N         Máximo de palabras del título en el nombre (defecto: 8)
    --no-skip-renamed     Procesa también PDFs que ya tienen formato año_autor_...
    --email EMAIL         Email para la API de Crossref (educado identificarse)

Variables de entorno (config/.env):
    CROSSREF_EMAIL        Email para las peticiones a Crossref (opcional)

Dependencias:
    pymupdf (fitz), requests, openpyxl, python-dotenv

Notas:
    - Desde 2026-05-28: guiones y espacios en títulos se convierten en '_'
      uniformemente (fix en shorten_title y sanitize_filename).
    - PDFs renombrados antes del fix pueden tener guiones en el stem;
      usar 6_Mantenimiento → Coherencia PDF/MD para detectar artefactos huérfanos.
"""

from __future__ import annotations

import argparse
import csv
import logging
import re
import sys
import time
import unicodedata
from pathlib import Path
from typing import Optional

import fitz  # PyMuPDF
import requests

sys.path.insert(0, str(Path(__file__).parent))
from utils.pdf_utils import extract_doi_from_pdf, extract_doi_from_text

# ============================================================
# CONFIGURACIÓN POR DEFECTO
# ============================================================

DEFAULT_FOLDER = Path("/Volumes/research/inbox")
DEFAULT_CSV = Path("/Volumes/research/metadatos/renombrado.csv")
DEFAULT_EMAIL = "martin.ramirez@uca.es"
DEFAULT_DOI_MANUAL = Path("/Volumes/research/metadatos/doi_manual.xlsx")
FAILED_SUBFOLDER_NAME = "fallidos"

log = logging.getLogger(__name__)

CROSSREF_API = "https://api.crossref.org/works/"
STOPWORDS = {
    "a", "an", "and", "as", "at", "by", "for", "from", "in", "into", "of",
    "on", "or", "the", "to", "with", "without", "via", "using", "use",
    "towards", "toward", "over", "under", "through", "during", "between",
    "de", "del", "la", "el", "los", "las", "y", "en", "por", "para", "con",
    "sin", "sobre", "un", "una", "unos", "unas"
}


# ============================================================
# UTILIDADES DE TEXTO
# ============================================================

def strip_accents(text: str) -> str:
    text = unicodedata.normalize("NFKD", text)
    return "".join(c for c in text if not unicodedata.combining(c))


def slugify(text: str) -> str:
    text = strip_accents(text).lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[''`´]", "", text)
    text = re.sub(r"[^a-z0-9\s\-]", " ", text)
    text = re.sub(r"[\s\-]+", "_", text.strip())
    text = re.sub(r"_+", "_", text)
    return text.strip("_")


def shorten_title(title: str, max_words: int = 8) -> str:
    clean = strip_accents(title).lower()
    clean = re.sub(r"[^a-z0-9\s]", " ", clean)
    words = [w for w in clean.split() if w and w not in STOPWORDS]
    if not words:
        words = clean.split()
    return "_".join(words[:max_words]) if words else "untitled"


def sanitize_filename(name: str, max_len: int = 180) -> str:
    name = re.sub(r'[<>:"/\\|?*]', "_", name)
    name = re.sub(r"\s+", "_", name)
    name = re.sub(r"_+", "_", name).strip("._ ")
    if len(name) > max_len:
        stem = Path(name).stem[: max_len - 4]
        name = f"{stem}.pdf"
    return name


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    i = 2
    while True:
        candidate = parent / f"{stem}_{i}{suffix}"
        if not candidate.exists():
            return candidate
        i += 1


def already_renamed(filename: str) -> bool:
    return bool(re.match(r"^\d{4}_[a-z0-9]+_.+\.pdf$", filename.lower()))


# ============================================================
# DOI MANUAL DESDE EXCEL
# ============================================================

def load_doi_manual(xlsx_path: Path) -> dict[str, str]:
    """Carga (y crea si no existe) el Excel de DOIs manuales.

    Devuelve {nombre_archivo: doi} solo para filas con DOI relleno válido.
    Si el fichero no existe lo crea con cabecera y devuelve {}.
    """
    import openpyxl

    if not xlsx_path.exists():
        try:
            xlsx_path.parent.mkdir(parents=True, exist_ok=True)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["nombre_archivo", "doi"])
            wb.save(xlsx_path)
            print(f"Excel DOI manual creado: {xlsx_path}")
        except Exception as e:
            print(f"Advertencia: no se pudo crear {xlsx_path}: {e}")
        return {}

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        mapping: dict[str, str] = {}
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row or len(row) < 2:
                continue
            filename = str(row[0]).strip() if row[0] is not None else ""
            doi = str(row[1]).strip() if row[1] is not None else ""
            if filename and doi.lower().startswith("10."):
                mapping[filename] = doi
        wb.close()
        return mapping
    except Exception as e:
        print(f"Advertencia: no se pudo leer {xlsx_path}: {e}")
        return {}


def append_doi_not_found(xlsx_path: Path, filenames: list[str]) -> int:
    """Añade al Excel una fila vacía de DOI por cada nombre no presente aún.

    No sobreescribe filas existentes (tengan DOI o no).
    Devuelve el número de filas añadidas.
    """
    if not filenames:
        return 0
    if not xlsx_path.exists():
        return 0
    import openpyxl
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        existing: set[str] = set()
        if ws.max_row is not None and ws.max_row >= 2:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] is not None:
                    existing.add(str(row[0]).strip())
        added = 0
        for name in filenames:
            if name not in existing:
                ws.append([name, ""])
                existing.add(name)
                added += 1
        if added:
            wb.save(xlsx_path)
        return added
    except Exception as e:
        log.warning("append_doi_not_found: %s: %s", type(e).__name__, e)
        return 0


def update_renamed_in_excel(
    xlsx_path: Path,
    renamed: list[dict[str, str]],
) -> int:
    """Actualiza el Excel cuando un PDF se renombra con éxito.

    Para cada entrada {"original": str, "nuevo": str, "doi": str}:
    - Si nombre_original existe en col A → reemplaza por nombre_nuevo (y rellena DOI si vacío).
    - Si ninguno existe → añade fila nueva con (nombre_nuevo, doi).
    - Si nombre_nuevo ya existe → no hace nada (evita duplicados).
    Devuelve el número de filas modificadas o añadidas.
    """
    if not renamed:
        return 0
    if not xlsx_path.exists():
        return 0
    import openpyxl
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active

        # Índice nombre → número de fila (1-based), desde la fila 2
        name_to_row: dict[str, int] = {}
        if ws.max_row is not None and ws.max_row >= 2:
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=1).value
                if val is not None:
                    name_to_row[str(val).strip()] = row_idx

        changed = 0
        for entry in renamed:
            original = entry["original"]
            nuevo = entry["nuevo"]
            doi = entry["doi"]

            if nuevo in name_to_row:
                continue
            elif original in name_to_row:
                row_idx = name_to_row[original]
                ws.cell(row=row_idx, column=1).value = nuevo
                if doi and not ws.cell(row=row_idx, column=2).value:
                    ws.cell(row=row_idx, column=2).value = doi
                name_to_row[nuevo] = name_to_row.pop(original)
                changed += 1
            else:
                ws.append([nuevo, doi])
                name_to_row[nuevo] = ws.max_row
                changed += 1

        if changed:
            wb.save(xlsx_path)
        return changed
    except Exception as e:
        log.warning("update_renamed_in_excel: %s: %s", type(e).__name__, e)
        return 0


# ============================================================
# CROSSREF
# ============================================================

def fetch_crossref_metadata(doi: str, email: str, timeout: int = 20) -> dict:
    url = CROSSREF_API + requests.utils.quote(doi, safe="")
    headers = {"User-Agent": f"PDFRenamer/1.0 (mailto:{email})"}
    params = {"mailto": email}
    r = requests.get(url, headers=headers, params=params, timeout=timeout)
    r.raise_for_status()
    data = r.json()
    return data["message"]


def extract_year(message: dict) -> str:
    for field in ("issued", "published-print", "published-online", "created"):
        value = message.get(field)
        if isinstance(value, dict):
            parts = value.get("date-parts")
            if parts and isinstance(parts, list) and parts[0]:
                year = parts[0][0]
                return str(year)
    return "0000"


def extract_first_author_surname(message: dict) -> str:
    authors = message.get("author", [])
    if not authors:
        return "unknownauthor"

    first = authors[0]
    family = first.get("family")
    if family:
        return slugify(family) or "unknownauthor"

    name = first.get("name", "")
    if name:
        return slugify(name.split()[-1]) or "unknownauthor"

    return "unknownauthor"


def extract_title(message: dict) -> str:
    titles = message.get("title", [])
    if isinstance(titles, list) and titles and titles[0].strip():
        return titles[0].strip()
    return "untitled"


def build_new_filename(message: dict, max_words: int = 8) -> str:
    year = extract_year(message)
    surname = extract_first_author_surname(message)
    title = extract_title(message)
    short_title = shorten_title(title, max_words=max_words)
    return sanitize_filename(f"{year}_{surname}_{short_title}.pdf")


# ============================================================
# FICHEROS FALLIDOS
# ============================================================

def ensure_failed_folder(folder: Path) -> Path:
    failed = folder / FAILED_SUBFOLDER_NAME
    failed.mkdir(parents=True, exist_ok=True)
    return failed


def move_to_failed(pdf_path: Path, failed_folder: Path) -> Path:
    target = unique_path(failed_folder / pdf_path.name)
    pdf_path.rename(target)
    return target


# ============================================================
# PROCESADO
# ============================================================

def process_pdf(
    pdf_path: Path,
    email: str,
    apply_changes: bool,
    max_words: int,
    skip_already_renamed: bool,
    failed_folder: Optional[Path],
    doi_manual: dict[str, str],
) -> dict:
    """Procesa un PDF. Si failed_folder es None, los fallidos no se mueven."""
    original = pdf_path.name
    result: dict = {
        "archivo_original": original,
        "doi": "",
        "year": "",
        "first_author": "",
        "title": "",
        "archivo_resultado": "",
        "estado": "",
        "error": "",
    }

    if skip_already_renamed and already_renamed(original):
        result["archivo_resultado"] = original
        result["estado"] = "IGNORADO_YA_RENOMBRADO"
        return result

    # Consultar DOI manual antes de intentar extraerlo del PDF
    manual_doi = doi_manual.get(original)
    if manual_doi:
        doi: Optional[str] = manual_doi
    else:
        # Extraer DOI (extract_doi_from_pdf ya atrapa sus propias excepciones)
        try:
            doi = extract_doi_from_pdf(pdf_path)
        except Exception as e:
            result["estado"] = "ERROR_LECTURA_PDF"
            result["error"] = f"{type(e).__name__}: {e}"
            if apply_changes and failed_folder is not None:
                moved = move_to_failed(pdf_path, failed_folder)
                result["archivo_resultado"] = moved.name
                result["estado"] = "MOVIDO_A_FALLIDOS_ERROR_LECTURA_PDF"
            return result

        if not doi:
            result["estado"] = "DOI_NO_ENCONTRADO"
            result["error"] = "No se encontró DOI en metadatos ni en el texto del PDF"
            if apply_changes and failed_folder is not None:
                moved = move_to_failed(pdf_path, failed_folder)
                result["archivo_resultado"] = moved.name
                result["estado"] = "MOVIDO_A_FALLIDOS_DOI_NO_ENCONTRADO"
            return result

    result["doi"] = doi

    try:
        metadata = fetch_crossref_metadata(doi, email=email)
        year = extract_year(metadata)
        first_author = extract_first_author_surname(metadata)
        title = extract_title(metadata)

        result["year"] = year
        result["first_author"] = first_author
        result["title"] = title

        new_name = build_new_filename(metadata, max_words=max_words)
        target = unique_path(pdf_path.with_name(new_name))

        if apply_changes:
            pdf_path.rename(target)
            result["archivo_resultado"] = target.name
            result["estado"] = "DOI_MANUAL" if manual_doi else "RENOMBRADO"
        else:
            result["archivo_resultado"] = target.name
            result["estado"] = "DOI_MANUAL" if manual_doi else "PREVIEW"

    except requests.HTTPError as e:
        status_code = e.response.status_code if e.response is not None else "NA"
        result["estado"] = f"HTTP_ERROR_{status_code}"
        result["error"] = str(e)
        if apply_changes and failed_folder is not None:
            moved = move_to_failed(pdf_path, failed_folder)
            result["archivo_resultado"] = moved.name
            result["estado"] = f"MOVIDO_A_FALLIDOS_HTTP_ERROR_{status_code}"

    except Exception as e:
        result["estado"] = f"ERROR_{type(e).__name__}"
        result["error"] = str(e)
        if apply_changes and failed_folder is not None:
            moved = move_to_failed(pdf_path, failed_folder)
            result["archivo_resultado"] = moved.name
            result["estado"] = f"MOVIDO_A_FALLIDOS_ERROR_{type(e).__name__}"

    return result


def resolve_csv_path(csv_arg: str, folder: Path) -> Path:
    """Resuelve la ruta del CSV:
    - Ruta absoluta → úsala tal cual.
    - Solo nombre de archivo → guárdalo dentro de la carpeta procesada.
    - Ruta relativa con directorio → resuélvela desde cwd.
    """
    p = Path(csv_arg)
    if p.is_absolute():
        return p
    if p.parent == Path("."):
        return folder / p
    return p.resolve()


def save_csv(rows: list[dict], csv_path: Path) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "archivo_original", "doi", "year", "first_author", "title",
        "archivo_resultado", "estado", "error",
    ]
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


# ============================================================
# MAIN
# ============================================================

def main() -> int:
    parser = argparse.ArgumentParser(description="Renombrar PDFs por DOI usando Crossref")
    parser.add_argument("--folder", default=str(DEFAULT_FOLDER), help="Carpeta con PDFs")
    parser.add_argument("--email", default=DEFAULT_EMAIL, help="Email para Crossref")
    parser.add_argument("--apply", action="store_true", help="Renombrar de verdad")
    parser.add_argument(
        "--move-failed", action="store_true",
        help="Mover PDFs sin DOI o con error a la subcarpeta 'fallidos'"
    )
    parser.add_argument("--max-words", type=int, default=8, help="Máximo de palabras del título")
    parser.add_argument("--csv", default=str(DEFAULT_CSV), help="Ruta del CSV resumen")
    parser.add_argument(
        "--no-skip-renamed", action="store_true",
        help="No ignorar archivos ya renombrados"
    )
    parser.add_argument(
        "--doi-manual", default=str(DEFAULT_DOI_MANUAL),
        help="Excel con DOIs manuales: columna A = nombre PDF, columna B = DOI"
    )
    args = parser.parse_args()

    xlsx_path = Path(args.doi_manual)
    doi_manual = load_doi_manual(xlsx_path)

    folder = Path(args.folder).expanduser().resolve()
    if not folder.exists() or not folder.is_dir():
        print(f"Carpeta no válida: {folder}")
        return 1

    failed_folder: Optional[Path] = None
    if args.move_failed:
        failed_folder = ensure_failed_folder(folder)

    pdfs = sorted(set(folder.glob("*.pdf")) | set(folder.glob("*.PDF")))
    if not pdfs:
        print(f"No se encontraron PDFs en {folder}")
        return 0

    csv_path = resolve_csv_path(args.csv, folder)

    print(f"\nCarpeta:     {folder}")
    print(f"PDFs:        {len(pdfs)}")
    print(f"Email:       {args.email}")
    print(f"Modo:        {'RENOMBRADO REAL' if args.apply else 'PREVIEW'}")
    print(f"Fallidos:    {'→ subcarpeta fallidos' if args.move_failed else 'se quedan en su lugar'}")
    print(f"DOI manual:  {len(doi_manual)} entradas ({args.doi_manual})")
    print(f"CSV salida:  {csv_path}")
    print("-" * 90)

    rows = []
    for pdf in pdfs:
        row = process_pdf(
            pdf_path=pdf,
            email=args.email,
            apply_changes=args.apply,
            max_words=args.max_words,
            skip_already_renamed=not args.no_skip_renamed,
            failed_folder=failed_folder,
            doi_manual=doi_manual,
        )
        rows.append(row)

        print(f"[{row['estado']}] {row['archivo_original']}")
        if row["doi"]:
            print(f"   DOI:       {row['doi']}")
        if row["year"] or row["first_author"]:
            print(f"   Autor/año: {row['first_author']} ({row['year']})")
        if row["archivo_resultado"] and row["archivo_resultado"] != row["archivo_original"]:
            print(f"   Resultado: {row['archivo_resultado']}")
        if row["error"]:
            print(f"   Error:     {row['error']}")
        print()

        time.sleep(0.2)

    save_csv(rows, csv_path)

    not_found = [
        r["archivo_original"] for r in rows
        if r["estado"] in {"DOI_NO_ENCONTRADO", "MOVIDO_A_FALLIDOS_DOI_NO_ENCONTRADO"}
    ]
    added = append_doi_not_found(xlsx_path, not_found)
    if added:
        print(f"Excel DOI manual: {added} fila(s) nueva(s) añadida(s) → {xlsx_path}")

    if args.apply:
        renamed_entries = [
            {"original": r["archivo_original"], "nuevo": r["archivo_resultado"], "doi": r["doi"]}
            for r in rows
            if r["estado"] in {"RENOMBRADO", "DOI_MANUAL"}
        ]
        changed = update_renamed_in_excel(xlsx_path, renamed_entries)
        if changed:
            print(f"Excel DOI manual: {changed} entrada(s) actualizadas tras renombrado → {xlsx_path}")

    ok = sum(1 for r in rows if r["estado"] in {"PREVIEW", "RENOMBRADO", "DOI_MANUAL"})
    moved = sum(1 for r in rows if "MOVIDO_A_FALLIDOS" in r["estado"])
    ignored = sum(1 for r in rows if r["estado"] == "IGNORADO_YA_RENOMBRADO")
    failed_count = len(rows) - ok - moved - ignored

    print("-" * 90)
    print(f"Correctos:                {ok}")
    print(f"Movidos a fallidos:       {moved}")
    print(f"Ignorados ya renombrados: {ignored}")
    print(f"Incidencias:              {failed_count}")
    print(f"CSV guardado en:          {csv_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
